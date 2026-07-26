from __future__ import annotations

import csv
from pathlib import Path
from typing import TypeAlias

from openpyxl import load_workbook

from ..database import Database
from ..models import BatchPreviewRow, LoopSettings, ProcessingSettings
from ..paths import normalize_internal_path

REQUIRED_COLUMNS = {"original_path", "replacement_file"}
ResolvedBatchRow: TypeAlias = tuple[BatchPreviewRow, Path | None]

PROCESSING_COLUMNS = {
    "trim_start": "trim_start_seconds",
    "trim_start_seconds": "trim_start_seconds",
    "trim_end": "trim_end_seconds",
    "trim_end_seconds": "trim_end_seconds",
    "fade_in": "fade_in_seconds",
    "fade_in_seconds": "fade_in_seconds",
    "fade_out": "fade_out_seconds",
    "fade_out_seconds": "fade_out_seconds",
    "gain_db": "gain_db",
    "normalize": "normalize",
    "auto_trim_silence": "auto_trim_silence",
    "sample_rate": "sample_rate",
    "channels": "channels",
}
LOOP_COLUMNS = {
    "loop_enabled": "enabled",
    "loop_start": "start_seconds",
    "loop_start_seconds": "start_seconds",
    "loop_end": "end_seconds",
    "loop_end_seconds": "end_seconds",
}


def _boolean(value: object) -> bool:
    normalized = str(value).strip().casefold()
    if normalized in {"1", "true", "yes", "on"}:
        return True
    if normalized in {"0", "false", "no", "off"}:
        return False
    raise ValueError(f"Expected a boolean value, received {value!r}.")


def _row_settings(
    row: dict[str, object],
) -> tuple[ProcessingSettings, LoopSettings, bool]:
    processing: dict[str, object] = {}
    looping: dict[str, object] = {}
    uses_row_settings = False
    for column, field in PROCESSING_COLUMNS.items():
        raw = row.get(column)
        if raw is None or str(raw).strip() == "":
            continue
        uses_row_settings = True
        if field in {"normalize", "auto_trim_silence"}:
            processing[field] = _boolean(raw)
        elif field in {"sample_rate", "channels"}:
            processing[field] = int(raw)
        else:
            processing[field] = float(raw)
    for column, field in LOOP_COLUMNS.items():
        raw = row.get(column)
        if raw is None or str(raw).strip() == "":
            continue
        uses_row_settings = True
        if field == "enabled":
            looping[field] = _boolean(raw)
        else:
            looping[field] = float(raw)
    return (
        ProcessingSettings.model_validate(processing),
        LoopSettings.model_validate(looping),
        uses_row_settings,
    )


def preview_mapping_file(
    database: Database, mapping_file: Path
) -> list[BatchPreviewRow]:
    return [row for row, _ in resolve_mapping_file(database, mapping_file)]


def resolve_mapping_file(
    database: Database, mapping_file: Path
) -> list[ResolvedBatchRow]:
    """Validate mapping rows and resolve each unambiguous source/target pair."""
    path = mapping_file.resolve(strict=True)
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as stream:
            rows = list(csv.DictReader(stream))
            fieldnames = set()
            if rows:
                fieldnames = set(rows[0].keys())
    elif path.suffix.lower() == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
        headers = []
        if values:
            headers = [str(value or "").strip() for value in values[0]]
        rows = [
            dict(zip(headers, values_row, strict=False))
            for values_row in values[1:]
        ]
        fieldnames = set(headers)
    else:
        return [
            (
                BatchPreviewRow(
                    row_number=1,
                    original_path="",
                    replacement_file="",
                    status="unsupportedFile",
                    messages=["Choose a CSV or XLSX mapping file."],
                ),
                None,
            )
        ]
    missing = REQUIRED_COLUMNS - fieldnames
    if missing:
        return [
            (
                BatchPreviewRow(
                    row_number=1,
                    original_path="",
                    replacement_file="",
                    status="invalid",
                    messages=[f"Missing columns: {', '.join(sorted(missing))}"],
                ),
                None,
            )
        ]
    result: list[ResolvedBatchRow] = []
    for number, row in enumerate(rows, start=2):
        original = str(row.get("original_path") or "").strip()
        replacement = str(row.get("replacement_file") or "").strip()
        messages: list[str] = []
        asset = None
        processing = ProcessingSettings()
        looping = LoopSettings()
        uses_row_settings = False
        try:
            processing, looping, uses_row_settings = _row_settings(row)
        except Exception as error:
            messages.append(f"Invalid per-row processing settings: {error}")
        try:
            normalized = normalize_internal_path(original)
            asset = database.get_asset_by_path(normalized)
        except Exception as error:
            messages.append(str(error))
        raw_source = Path(replacement)
        if raw_source.is_absolute():
            source = raw_source.resolve(strict=False)
        else:
            source = (path.parent / raw_source).resolve(strict=False)
        if not raw_source.is_absolute():
            try:
                source.relative_to(path.parent.resolve())
            except ValueError:
                messages.append("Relative replacement path escapes the mapping folder.")
        if not asset:
            messages.append("No indexed target matches original_path.")
        if not source.is_file():
            messages.append("Replacement file is missing.")
        if source.suffix.lower() not in {".wav", ".mp3"}:
            messages.append("Replacement must be MP3 or WAV.")
        asset_id = None
        if asset:
            asset_id = asset.id
        status = "matched"
        resolved_source = source
        if messages:
            status = "invalid"
            resolved_source = None
        preview = BatchPreviewRow(
            row_number=number,
            original_path=original,
            replacement_file=replacement,
            asset_id=asset_id,
            status=status,
            messages=messages,
            processing=processing,
            looping=looping,
            uses_row_settings=uses_row_settings,
        )
        result.append((preview, resolved_source))
    return result


def preview_folder(database: Database, folder: Path) -> list[BatchPreviewRow]:
    return [row for row, _ in resolve_folder(database, folder)]


def resolve_folder(database: Database, folder: Path) -> list[ResolvedBatchRow]:
    """Match audio filenames in a folder against the indexed catalog."""
    root = folder.resolve(strict=True)
    rows: list[ResolvedBatchRow] = []
    sources = sorted(
        path
        for path in root.rglob("*")
        if path.is_file() and path.suffix.lower() in {".wav", ".mp3"}
    )
    for number, source in enumerate(sources, start=1):
        candidates = database.get_assets_by_filename(f"{source.stem}.vsnd_c")
        status = "missingTarget"
        original_path = ""
        asset_id = None
        messages = ["No filename match."]
        resolved_source = None
        if len(candidates) == 1:
            status = "matched"
            original_path = candidates[0].internal_path
            asset_id = candidates[0].id
            messages = []
            resolved_source = source
        elif candidates:
            status = "ambiguous"
            messages = [f"{len(candidates)} possible targets."]
        preview = BatchPreviewRow(
            row_number=number,
            original_path=original_path,
            replacement_file=str(source.relative_to(root)),
            asset_id=asset_id,
            status=status,
            messages=messages,
        )
        rows.append((preview, resolved_source))
    return rows
